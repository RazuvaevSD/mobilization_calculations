from os import path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.services.db.document import Document

EXP_FACTOR = 1.2  # расширяющий коэфициент для не моно шрифтов
SPARE = 2  # запас для отступаов до края ячейки
FONT_SIZE = 12
FONT = Font(size=FONT_SIZE, name='Times New Roman')

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

ANNEX_TEXT = [
    '',
    'Приложение: 1. Объем бюджетных ассигнований, необходимых для выполнения '
    '\n                         '
    'мероприятий Плана перевода ОГВ на работу в условиях военного времени.',
    'Приложение: 1. Объем бюджетных ассигнований ОГВ на выполнение '
    'мобилизационных'
    '\n                         '
    'заданий (задач), установленных Мобилизационным планом экономики'
    '\n                         '
    'Сахалинской области на расчетный период военного времени.',
    'Приложение: 1. Объем бюджетных ассигнований, необходимых для выполнения '
    '\n                         '
    'мероприятий Плана перевода ОГВ на работу в условиях военного времени.'
    '\n                         '
    '2. Объем бюджетных ассигнований ОГВ на выполнение мобилизационных'
    '\n                         '
    'заданий (задач), установленных Мобилизационным планом экономики'
    '\n                         '
    'Сахалинской области на расчетный период военного времени.',
]


class Excel:
    @classmethod
    def write(cls, doc_ids, folder):
        for id in doc_ids:
            doc_header_data, total_amount, amount_detail = (
                Document.get_by_id(id)
            )
            filename = path.join(folder,
                                 f'{id} {doc_header_data.date}.xlsx')
            # создать книгу
            book = Workbook()
            # получить активный лист (автоматически создан)
            sheet = book.active
            # Заголовок
            sheet['E1'] = 'Для служебного использования'
            sheet.merge_cells('E1:F1')
            sheet['E1'].alignment = Alignment(horizontal='right')

            sheet['E2'] = 'Экз. № ' + doc_header_data.number or '1'
            sheet.merge_cells('E2:F2')
            sheet['E2'].alignment = Alignment(horizontal='right')

            sheet['A4'] = 'РАСЧЕТ ОБЪЕМА'
            sheet.merge_cells('A4:F4')
            sheet['A4'].alignment = Alignment(horizontal='center')
            if doc_header_data.doc_type == 'amount_oiv':
                sheet['A5'] = ('расходных обязательств, необходимых\n' +
                               doc_header_data.org_name or '')
                sheet.merge_cells('A5:F5')
                sheet['A5'].alignment = Alignment(horizontal='center',
                                                  wrap_text=True)
                sheet.row_dimensions[5].height = cls._count_height(2)

                sheet['A6'] = 'для выполнения мероприятий по мобилизации'
                sheet.merge_cells('A6:F6')
                sheet['A6'].alignment = Alignment(horizontal='center',
                                                  wrap_text=True)
            else:
                sheet['A5'] = ('расходных обязательств, необходимых')
                sheet.merge_cells('A5:F5')
                sheet['A5'].alignment = Alignment(horizontal='center',
                                                  wrap_text=True)

                sheet['A6'] = ('для выполнения мероприятий по мобилизации'
                               ' в Сахалинской области')
                sheet.merge_cells('A6:F6')
                sheet['A6'].alignment = Alignment(horizontal='center',
                                                  wrap_text=True)

            # Заголовок таблицы
            sheet['F8'] = 'тыс.руб.'
            sheet['F8'].alignment = Alignment(horizontal='right')
            sheet['A9'] = '№\nп/п'
            sheet.merge_cells('A9:A10')
            sheet['B9'] = 'Наименование'
            sheet.merge_cells('B9:B10')
            sheet['C9'] = 'ВР\n(группа,\nподгруппа,\nэлемент)'
            sheet.merge_cells('C9:C10')
            sheet['D9'] = 'Объем бюджетных ассигнований,\nв том числе:'
            sheet.merge_cells('D9:E9')
            sheet['D10'] = ('на выполнение\nмероприятий\nПлана перевода\n'
                            'на работу в\nусловиях\nвоенного\nвремени')
            sheet['E10'] = (
                'на выполнение\nмобилизационных\nзаданий (задач),\n'
                'установленных\nМобилизационным\nпланом\nэкономики\n'
                'Сахалинской\nобласти на\nрасчетный\nпериод\n'
                'военного\nвремени')
            sheet['F9'] = 'Всего'
            sheet.merge_cells('F9:F10')
            sheet['A11'] = '1'
            sheet['B11'] = '2'
            sheet['C11'] = '3'
            sheet['D11'] = '4'
            sheet['E11'] = '5'
            sheet['F11'] = '6'
            # выравнивание заголовка таблицы
            for row in sheet.iter_rows(min_row=9, max_row=11, max_col=6):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='top',
                                               wrap_text=True)
            # высота строки = размер шрифта * количество подстрок
            #                 * коэфициент (типа межстрочного интервала)
            sheet.row_dimensions[9].height = cls._count_height(2)
            sheet.row_dimensions[10].height = cls._count_height(13)

            sheet.column_dimensions['A'].width = cls._count_whidth(2)
            sheet.column_dimensions['B'].width = cls._count_whidth(19)
            sheet.column_dimensions['C'].width = cls._count_whidth(8)
            sheet.column_dimensions['D'].width = cls._count_whidth(15)
            sheet.column_dimensions['E'].width = cls._count_whidth(15)
            sheet.column_dimensions['F'].width = cls._count_whidth(6)

            # Итоговая строка
            sheet['A12'] = ''
            sheet['B12'] = total_amount.name
            sheet['D12'] = int(total_amount.amounts_transfer)
            sheet['E12'] = int(total_amount.amount_economy)
            sheet['F12'] = int(total_amount.total)
            # Детализированные показатели
            row_index = 12
            for amount in amount_detail:
                row_index += 1
                sheet[f'A{row_index}'] = row_index - 12
                sheet[f'C{row_index}'].alignment = Alignment(
                    horizontal='center')
                sheet[f'B{row_index}'] = amount.name
                sheet[f'C{row_index}'].alignment = Alignment(
                    horizontal='left',
                    wrap_text=True)
                sheet[f'C{row_index}'] = int(amount.type_expenses)
                sheet[f'C{row_index}'].alignment = Alignment(
                    horizontal='center')
                sheet[f'D{row_index}'] = int(amount.amounts_transfer)
                sheet[f'E{row_index}'] = int(amount.amount_economy)
                sheet[f'F{row_index}'] = int(amount.total)

                sheet[f'B{row_index}'].alignment = Alignment(wrap_text=True)
                # Высота строки = длина значения //
                #                 (ширину колонки / коэфициент расширения -
                #                  запас) + 1
                subrow_count = int(len(sheet[f'B{row_index}'].value) //
                                   (sheet.column_dimensions['B'].width /
                                    EXP_FACTOR - SPARE)) + 1
                sheet.row_dimensions[row_index].height = cls._count_height(
                    subrow_count)

            # Рисуем рамку для таблицы
            for row in sheet.iter_rows(min_row=9, max_row=row_index,
                                       max_col=6):
                for cell in row:
                    cell.border = THIN_BORDER

            # Приложение
            if doc_header_data.annex != '0':
                row_index += 2
                sheet[f'A{row_index}'] = ANNEX_TEXT[int(doc_header_data.annex)]
                sheet.merge_cells(f'A{row_index}:F{row_index}')
                sheet[f'A{row_index}'].alignment = Alignment(horizontal='left',
                                                             wrap_text=True)
                subrow_count = int(len(sheet[f'A{row_index}'].value) //
                                   ((sheet.column_dimensions['A'].width +
                                     sheet.column_dimensions['B'].width +
                                     sheet.column_dimensions['C'].width +
                                     sheet.column_dimensions['D'].width +
                                     sheet.column_dimensions['E'].width +
                                     sheet.column_dimensions['F'].width) /
                                    EXP_FACTOR - SPARE)) + 1
                print('len', len(sheet[f'A{row_index}'].value),
                      'width', sheet.column_dimensions['A'].width,
                      'factor', EXP_FACTOR,
                      'spare', SPARE,
                      subrow_count)
                sheet.row_dimensions[row_index].height = cls._count_height(
                    subrow_count)
            # Подпись руководителя ОИВ
            row_index += 2
            sheet[f'A{row_index}'] = doc_header_data.pos_head
            sheet.merge_cells(f'A{row_index}:D{row_index}')
            sheet[f'A{row_index}'].alignment = Alignment(horizontal='left',
                                                         wrap_text=True)
            sheet[f'E{row_index}'] = '_' * 20
            sheet.merge_cells(f'E{row_index}:F{row_index}')
            sheet[f'A{row_index}'].alignment = Alignment(horizontal='left',
                                                         vertical='bottom',
                                                         wrap_text=True)
            # Подпись согласующего
            row_index += 2
            sheet[f'A{row_index}'] = 'Согласовано:'
            row_index += 1
            sheet[f'A{row_index}'] = doc_header_data.pos_coordinator
            sheet.merge_cells(f'A{row_index}:D{row_index}')
            sheet[f'A{row_index}'].alignment = Alignment(horizontal='left',
                                                         wrap_text=True)
            sheet[f'E{row_index}'] = '_' * 20
            sheet.merge_cells(f'E{row_index}:F{row_index}')
            sheet[f'A{row_index}'].alignment = Alignment(horizontal='left',
                                                         vertical='bottom',
                                                         wrap_text=True)

            # Установить шрифт для всех ячеек
            for row in sheet.iter_rows(min_row=1, max_row=row_index,
                                       max_col=6):
                for cell in row:
                    cell.font = FONT

            book.save(filename)

    @staticmethod
    def _count_whidth(value_len):
        # Ширина колонки = (самая длинная подстрока + запас для отступов)
        #                  * расширяющий коэфициент для не моно шрифтов
        return (value_len + SPARE) * EXP_FACTOR

    @staticmethod
    def _count_height(subrow_count):
        # высота строки = размер шрифта * количество подстрок
        #                 * коэфициент (типа межстрочного интервала)
        return FONT_SIZE * subrow_count * EXP_FACTOR
